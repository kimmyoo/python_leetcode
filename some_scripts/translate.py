import json
from openpyxl import load_workbook

def translate_excel(input_excel, json_file):
    # Load the JSON dictionary
    with open(json_file, 'r', encoding='utf-8') as f:
        dictionary = json.load(f)
    # Load the Excel workbook
    workbook = load_workbook(input_excel)
    sheet = workbook.active
    # List to store translated values
    translated_values = []
    
    # Iterate through each cell in the first column
    for row in sheet.iter_rows(min_row=1, 
                               max_col=1, 
                               max_row=sheet.max_row
                            ):
        # Store the original value of the cell
        original_value = row[0].value  
        # Initialize translated value which is the original string but in lowercase
        translated_value = original_value.lower()
        # Iterate through each key in the JSON dictionary
        for key in dictionary:
            # still need to use lowercase for checking
            if key in original_value.lower():
                # Replace the key with its corresponding value
                translated_value = translated_value.replace(key, dictionary[key])
        # Append translated value to list
        translated_values.append(translated_value)

    # Writing the translated values to the next column
    for idx, value in enumerate(translated_values, start=1):
        translated_cell = sheet.cell(row=idx, column=2)
        translated_cell.value = value

    # Save workbook
    workbook.save(input_excel)

if __name__ == "__main__":
    input_excel = "target.xlsx"
    json_file = "dict.json"

    translate_excel(input_excel, json_file)
